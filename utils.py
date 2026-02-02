import csv
import openpyxl
import json
from re import compile, match
import json
from time import sleep
from random import uniform
from datetime import datetime, timedelta
from typing import List, Dict
from collections import defaultdict


def overdue_date(expire_date_str: str):
    date_format = "%m/%d/%y %I:%M %p"
    expire_date = datetime.strptime(expire_date_str, date_format)
    now = datetime.now()
    delta = expire_date - now
    return f"{delta}"


def cleared_date() -> str:
    date_format = "%m/%d/%y %I:%M %p"
    yesterday = datetime.now() - timedelta(days=1)
    return yesterday.strftime(date_format) 


def expire_date_days(expire_date_str: str):
    date_format = "%m/%d/%y %I:%M %p"
    expire_date = datetime.strptime(expire_date_str, date_format)
    now = datetime.now()
    delta = expire_date - now
    return delta.days


def group_keys_by_value(data: List[Dict[str, str]]) -> Dict[str, List[str]]:
    grouped = defaultdict(list)
    for item in data:
        for key, value in item.items():
            grouped[value].append(key)
    return dict(grouped)


def is_clear_or_marked(status: str) -> bool:
    is_clear_or_marked = compile(r"marked|clear")
    return True if is_clear_or_marked.match(status.lower()) else False


def get_not_completed_status_history(status_history: list[dict]) -> list[dict]:
    not_completed = []
    for sh in status_history:
        for s in sh.values():
            if not is_clear_or_marked(s):
                not_completed.append(sh)
    return not_completed


def is_completed_status_history(status_history: list[dict]) -> bool:
    status = []
    for sh in status_history:
        for s in sh.values():
            status.append(s)
    for s in status:
        if not is_clear_or_marked(s):
            return False
    return True


def delay(min: float, max: float) -> None:
    sleep(uniform(min, max))


def save_to_json(data: list[dict], output_filename: str ='output.json', indent: int = 2):
    try:
        with open(output_filename, 'w', encoding='utf-8') as f:
            json.dump(data, f, indent=indent, ensure_ascii=False)
        print(f"Successfully saved {len(data)} records to {output_filename}")
        return output_filename
    except Exception as e:
        print(f"Error saving to JSON: {e}")
        return None


def write_csv(filename: str, fields: list[str], data: list[dict]) -> None:
    if len(fields) > 0 and len(data) > 0:
        with open(filename, 'w', newline='') as f:
            writer = csv.DictWriter(f, fieldnames=fields)
            writer.writeheader()
            writer.writerows(data)


def read_json(filename: str) -> list[dict]:
    try:
        with open(filename, 'r') as file:
            data = json.load(file)
            if isinstance(data, list) and all(isinstance(item, dict) for item in data):
                return data
            else:
                raise ValueError("JSON file does not contain a list of dictionaries as its top-level structure.")
    except FileNotFoundError:
        print(f"Error: The file '{filename}' was not found.")
        return []
    except json.JSONDecodeError:
        print(f"Error: Could not decode JSON from the file '{filename}'. Check file format.")
        return []

def write_spreadsheet(filename: str, data: list[dict]) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "tickets"
    ws = wb["tickets"]

    i = 1
    cols = ["job_name", "cross_street", "ticket_type", "status", "id_ticket", "former_id_ticket", "release_date", "response_date", "expire_date", "permit", "days_to_expire", "cleared_ticket_date", "days_overdue"]
    for c in cols:
        ws.cell(1, i, c)
        i += 1

    row = 2
    for ticket in data:
        i = 1
        for v in ticket.values():
            ws.cell(row, i, v)
            i += 1
        row += 1

    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].bestFit = True

    #print(f"Writing to {filename}.")
    wb.save(filename)


def extract_without_parenthesis(text: str) -> str:
    pattern = r'^(.*?)\s*\(.*\).*$'
    res = match(pattern, text)
    if res:
        return res.group(1).strip()
    return text.strip()


def format_not_completed_status_history(status_history: list[dict]) -> str:
    group_by = group_keys_by_value(status_history)
    f = []
    for k, v in group_by.items():
        if len(v) > 1:
            s = f'{", ".join(v[:len(v)-1])} & {v[len(v)-1]}'
        else:
            s = f'{", ".join(v)}'
        e = extract_without_parenthesis(k)
        f.append(f"{s} - {e}")
    return " & ".join(f)


def check_ticket_type(former_id_ticker: str) -> str:
    return "New" if len(former_id_ticker) == 0 else "Update-Renewal"

